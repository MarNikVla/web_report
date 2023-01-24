from functools import lru_cache
import pathlib
from typing import Type

from django.core.files.storage import default_storage
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from web_report_card.grafik_via_classes import Worker as Wocker_grafik
from web_report_card.report_сard_via_classes import Worker as Wocker_table

WOCKER_TABLE = Wocker_table
WOCKER_GRAFIK = Wocker_grafik


@lru_cache
def get_workers(sheet: Type[Worksheet], start_row, final_row, col_of_names, woker=WOCKER_TABLE) -> \
        list[WOCKER_TABLE]:
    """
    get all worker from sheet
    :param sheet: Excel sheet
    :param start_row: START_ROW_OF_NAMES
    :param final_row: FINAL_ROW_OF_NAMES
    :param col_of_names: COLUMN_OF_NAMES
    :return: list of Worker instance
    """
    workers_list = list()

    # Iteration on column of names
    for col in sheet.iter_cols(min_row=start_row,
                               max_row=final_row,
                               min_col=col_of_names,
                               max_col=col_of_names):
        for cell in col:
            if cell.value is not None:
                workers_list.append(woker(cell.coordinate, sheet))
    return workers_list


def fill_all_workers(workers_list: list[WOCKER_TABLE]):
    """
    fill cells for all worker on sheet
    :param workers_list:
    :return: list[Worker]
    """
    for worker in workers_list:
        worker.fill_worker_line()


def make_backup(file_name):
    pass


def save_file(file_name):
    """
    save excel file with calculation results and do backup
    :param file_name: name of exel file
    :return:
    """

    start_row_of_names_grafik: int = 15
    final_row_of_names_grafik: int = 37
    column_of_names_grafik: int = 2

    report_card_file = pathlib.Path(file_name)
    backup_report_card_file = report_card_file.parent.joinpath(f'backup_{report_card_file.name}')
    wb = load_workbook(filename=report_card_file)
    for sheet in wb._sheets:
        fill_all_workers(get_workers(sheet, start_row_of_names_grafik, final_row_of_names_grafik,
                                     column_of_names_grafik, woker=WOCKER_GRAFIK))
    wb.save(report_card_file)


def make_table_file_for_web_app(file_name):
    """
    save excel file with populated results
    :param file_name: name of exel file
    :return: name of exel file
    """
    start_row_of_names_table: int = 15
    final_row_of_names_table: int = 49
    column_of_names_table: int = 2

    report_card_file_path = default_storage.path(file_name)
    wb = load_workbook(filename=report_card_file_path)
    for sheet in wb._sheets:
        fill_all_workers(get_workers(sheet, start_row_of_names_table, final_row_of_names_table,
                                     column_of_names_table, woker=WOCKER_TABLE))
    wb.save(report_card_file_path)
    return report_card_file_path


def make_grafik_file_for_web_app(file_name):
    """
    save excel file with populated results
    :param file_name: name of exel file
    :return: name of exel file
    """

    start_row_of_names_grafik: int = 15
    final_row_of_names_grafik: int = 37
    column_of_names_grafik: int = 2

    report_card_file_path = default_storage.path(file_name)
    wb = load_workbook(filename=report_card_file_path)
    for sheet in wb._sheets:
        fill_all_workers(get_workers(sheet, start_row_of_names_grafik, final_row_of_names_grafik,
                                     column_of_names_grafik, woker=WOCKER_GRAFIK))
    wb.save(report_card_file_path)
    return report_card_file_path
